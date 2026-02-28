import io
from django.http import HttpResponse
from django.shortcuts import render
from django.views.generic import View
from django.db.models import Count, Q

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from accounts.mixins import AdminRequiredMixin
from applications.models import Application
from companies.models import Company, JobPost
from students.models import StudentProfile


# ════════════════════════════════════════════════
#  HELPER: Style Functions
# ════════════════════════════════════════════════

def get_header_style():
    return {
        'font':      Font(bold=True, color='FFFFFF', size=11),
        'fill':      PatternFill('solid', fgColor='1a56db'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border':    Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
    }

def get_subheader_style():
    return {
        'font':      Font(bold=True, size=10),
        'fill':      PatternFill('solid', fgColor='E8F0FE'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border':    Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
    }

def apply_style(cell, style):
    cell.font      = style['font']
    cell.fill      = style['fill']
    cell.alignment = style['alignment']
    cell.border    = style['border']

def set_column_widths(ws, widths):
    for col_num, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width


# ════════════════════════════════════════════════
#  REPORTS HOME PAGE
# ════════════════════════════════════════════════

class ReportsHomeView(AdminRequiredMixin, View):

    def get(self, request):
        context = {
            'total_students':  StudentProfile.objects.count(),
            'total_placed':    StudentProfile.objects.filter(placement_status='placed').count(),
            'total_companies': Company.objects.filter(is_active=True).count(),
            'total_apps':      Application.objects.count(),
        }
        return render(request, 'reports/reports_home.html', context)


# ════════════════════════════════════════════════
#  REPORT 1: All Students
# ════════════════════════════════════════════════

class ExportStudentsView(AdminRequiredMixin, View):

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Students'

        # ── Title Row ───────────────────────────
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value     = 'ALL STUDENTS REPORT — Placement Portal'
        title_cell.font      = Font(bold=True, size=14, color='1a56db')
        title_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 30

        # ── Headers ─────────────────────────────
        headers = [
            'Roll No', 'Full Name', 'Email', 'Branch',
            'Year', 'CGPA', 'Backlogs', 'Phone',
            'Placement Status', 'Resume'
        ]
        widths = [14, 22, 28, 28, 10, 8, 10, 15, 18, 10]

        header_style = get_header_style()
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            apply_style(cell, header_style)

        set_column_widths(ws, widths)
        ws.row_dimensions[2].height = 20

        # ── Data Rows ────────────────────────────
        students = StudentProfile.objects.select_related('user').order_by('branch', 'roll_number')
        for row_num, student in enumerate(students, 3):
            row_data = [
                student.roll_number,
                student.user.full_name,
                student.user.email,
                student.get_branch_display(),
                f'Year {student.year}',
                float(student.cgpa),
                student.backlogs,
                student.phone or '—',
                student.get_placement_status_display(),
                'Yes' if student.resume else 'No',
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center')
                cell.border    = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'),
                )
                # Color placed rows green
                if student.placement_status == 'placed':
                    cell.fill = PatternFill('solid', fgColor='E8F5E9')

        # ── Freeze top rows ──────────────────────
        ws.freeze_panes = 'A3'

        # ── Response ────────────────────────────
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="students_report.xlsx"'
        wb.save(response)
        return response


# ════════════════════════════════════════════════
#  REPORT 2: Placed Students
# ════════════════════════════════════════════════

class ExportPlacedStudentsView(AdminRequiredMixin, View):

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Placed Students'

        # ── Title ────────────────────────────────
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value     = 'PLACED STUDENTS REPORT — Placement Portal'
        title_cell.font      = Font(bold=True, size=14, color='1a56db')
        title_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 30

        # ── Headers ─────────────────────────────
        headers = [
            'Roll No', 'Full Name', 'Branch',
            'CGPA', 'Company', 'Job Title', 'CTC', 'Type'
        ]
        widths = [14, 22, 28, 8, 24, 24, 14, 14]

        header_style = get_header_style()
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            apply_style(cell, header_style)

        set_column_widths(ws, widths)

        # ── Data ─────────────────────────────────
        placed_apps = Application.objects.filter(
            status='selected'
        ).select_related(
            'student__user', 'job__company'
        ).order_by('student__branch', 'student__roll_number')

        for row_num, app in enumerate(placed_apps, 3):
            row_data = [
                app.student.roll_number,
                app.student.user.full_name,
                app.student.get_branch_display(),
                float(app.student.cgpa),
                app.job.company.name,
                app.job.title,
                f'Rs.{app.job.ctc_min} - Rs.{app.job.ctc_max} LPA',
                app.job.get_job_type_display(),
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center')
                cell.fill      = PatternFill('solid', fgColor='E8F5E9')
                cell.border    = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'),
                )

        ws.freeze_panes = 'A3'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="placed_students.xlsx"'
        wb.save(response)
        return response


# ════════════════════════════════════════════════
#  REPORT 3: Applications Report
# ════════════════════════════════════════════════

class ExportApplicationsView(AdminRequiredMixin, View):

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Applications'

        # ── Title ────────────────────────────────
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value     = 'ALL APPLICATIONS REPORT — Placement Portal'
        title_cell.font      = Font(bold=True, size=14, color='1a56db')
        title_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 30

        # ── Headers ─────────────────────────────
        headers = [
            'Roll No', 'Student Name', 'Branch',
            'Company', 'Job Title', 'Status',
            'Round', 'Applied Date', 'Remarks'
        ]
        widths = [14, 22, 24, 22, 24, 14, 18, 16, 30]

        header_style = get_header_style()
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            apply_style(cell, header_style)

        set_column_widths(ws, widths)

        # ── Status Color Map ─────────────────────
        status_colors = {
            'selected':    'C8E6C9',
            'rejected':    'FFCDD2',
            'shortlisted': 'B3E5FC',
            'on_hold':     'FFF9C4',
            'applied':     'F5F5F5',
        }

        # ── Data ─────────────────────────────────
        apps = Application.objects.select_related(
            'student__user', 'job__company'
        ).order_by('-applied_at')

        for row_num, app in enumerate(apps, 3):
            row_data = [
                app.student.roll_number,
                app.student.user.full_name,
                app.student.get_branch_display(),
                app.job.company.name,
                app.job.title,
                app.get_status_display(),
                app.get_current_round_display(),
                app.applied_at.strftime('%d %b %Y'),
                app.remarks or '—',
            ]
            color = status_colors.get(app.status, 'FFFFFF')
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center')
                cell.fill      = PatternFill('solid', fgColor=color)
                cell.border    = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'),
                )

        ws.freeze_panes = 'A3'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="applications_report.xlsx"'
        wb.save(response)
        return response


# ════════════════════════════════════════════════
#  REPORT 4: Branch-wise Summary
# ════════════════════════════════════════════════

class ExportBranchSummaryView(AdminRequiredMixin, View):

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Branch Summary'

        # ── Title ────────────────────────────────
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value     = 'BRANCH-WISE PLACEMENT SUMMARY'
        title_cell.font      = Font(bold=True, size=14, color='1a56db')
        title_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 30

        # ── Headers ─────────────────────────────
        headers = [
            'Branch', 'Total Students',
            'Placed', 'Not Placed',
            'Placement %', 'Total Applications'
        ]
        widths = [30, 16, 10, 12, 14, 18]

        header_style = get_header_style()
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            apply_style(cell, header_style)

        set_column_widths(ws, widths)

        # ── Data ─────────────────────────────────
        for row_num, (val, label) in enumerate(StudentProfile.BRANCH_CHOICES, 3):
            total  = StudentProfile.objects.filter(branch=val).count()
            placed = StudentProfile.objects.filter(branch=val, placement_status='placed').count()
            apps   = Application.objects.filter(student__branch=val).count()
            pct    = round((placed / total * 100) if total else 0, 1)

            row_data = [label, total, placed, total - placed, f'{pct}%', apps]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center')
                cell.border    = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'),
                )
                if col_num == 5:  # Placement %
                    if pct >= 70:
                        cell.fill = PatternFill('solid', fgColor='C8E6C9')
                    elif pct >= 40:
                        cell.fill = PatternFill('solid', fgColor='FFF9C4')
                    else:
                        cell.fill = PatternFill('solid', fgColor='FFCDD2')

        ws.freeze_panes = 'A3'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="branch_summary.xlsx"'
        wb.save(response)
        return response


# ════════════════════════════════════════════════
#  REPORT 5: Company-wise Summary
# ════════════════════════════════════════════════

class ExportCompanySummaryView(AdminRequiredMixin, View):

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Company Summary'

        # ── Title ────────────────────────────────
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value     = 'COMPANY-WISE PLACEMENT SUMMARY'
        title_cell.font      = Font(bold=True, size=14, color='1a56db')
        title_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 30

        # ── Headers ─────────────────────────────
        headers = [
            'Company', 'Industry', 'Total Jobs',
            'Total Applications', 'Shortlisted',
            'Selected', 'Rejected'
        ]
        widths = [26, 24, 12, 18, 14, 12, 12]

        header_style = get_header_style()
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            apply_style(cell, header_style)

        set_column_widths(ws, widths)

        # ── Data ─────────────────────────────────
        companies = Company.objects.filter(is_active=True).order_by('name')
        for row_num, company in enumerate(companies, 3):
            apps = Application.objects.filter(job__company=company)
            row_data = [
                company.name,
                company.get_industry_display(),
                JobPost.objects.filter(company=company).count(),
                apps.count(),
                apps.filter(status='shortlisted').count(),
                apps.filter(status='selected').count(),
                apps.filter(status='rejected').count(),
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center')
                cell.border    = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'),
                )
                if col_num == 6 and value > 0:
                    cell.fill = PatternFill('solid', fgColor='C8E6C9')

        ws.freeze_panes = 'A3'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="company_summary.xlsx"'
        wb.save(response)
        return response


# ════════════════════════════════════════════════
#  REPORT 6: Full Placement Report (Multi-Sheet)
# ════════════════════════════════════════════════

class ExportFullReportView(AdminRequiredMixin, View):

    def get(self, request):
        wb = openpyxl.Workbook()
        header_style    = get_header_style()
        subheader_style = get_subheader_style()

        # ════════════════════════════
        #  SHEET 1: Summary
        # ════════════════════════════
        ws1 = wb.active
        ws1.title = 'Summary'

        ws1.merge_cells('A1:D1')
        ws1['A1'].value     = 'PLACEMENT REPORT — FULL SUMMARY'
        ws1['A1'].font      = Font(bold=True, size=14, color='1a56db')
        ws1['A1'].alignment = Alignment(horizontal='center')
        ws1.row_dimensions[1].height = 30

        total_students = StudentProfile.objects.count()
        total_placed   = StudentProfile.objects.filter(placement_status='placed').count()
        pct            = round((total_placed / total_students * 100) if total_students else 0, 1)

        summary_data = [
            ('Total Students',    total_students),
            ('Total Placed',      total_placed),
            ('Total Not Placed',  total_students - total_placed),
            ('Placement %',       f'{pct}%'),
            ('Total Companies',   Company.objects.filter(is_active=True).count()),
            ('Total Jobs',        JobPost.objects.count()),
            ('Total Applications',Application.objects.count()),
        ]

        for i, (label, value) in enumerate(summary_data, 3):
            ws1.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=value)

        ws1.column_dimensions['A'].width = 24
        ws1.column_dimensions['B'].width = 16

        # ════════════════════════════
        #  SHEET 2: Placed Students
        # ════════════════════════════
        ws2 = wb.create_sheet('Placed Students')
        ws2.merge_cells('A1:H1')
        ws2['A1'].value     = 'PLACED STUDENTS'
        ws2['A1'].font      = Font(bold=True, size=13, color='1a56db')
        ws2['A1'].alignment = Alignment(horizontal='center')

        headers2 = [
            'Roll No', 'Name', 'Branch', 'CGPA',
            'Company', 'Job Title', 'CTC', 'Type'
        ]
        for col, h in enumerate(headers2, 1):
            apply_style(ws2.cell(row=2, column=col, value=h), header_style)

        set_column_widths(ws2, [14, 22, 24, 8, 24, 24, 18, 14])

        placed_apps = Application.objects.filter(
            status='selected'
        ).select_related('student__user', 'job__company')
        for row_num, app in enumerate(placed_apps, 3):
            for col, val in enumerate([
                app.student.roll_number,
                app.student.user.full_name,
                app.student.get_branch_display(),
                float(app.student.cgpa),
                app.job.company.name,
                app.job.title,
                f'Rs.{app.job.ctc_min} - Rs.{app.job.ctc_max} LPA',
                app.job.get_job_type_display(),
            ], 1):
                cell = ws2.cell(row=row_num, column=col, value=val)
                cell.alignment = Alignment(horizontal='center')
                cell.fill      = PatternFill('solid', fgColor='E8F5E9')
                cell.border    = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'),
                )

        # ════════════════════════════
        #  SHEET 3: Branch Summary
        # ════════════════════════════
        ws3 = wb.create_sheet('Branch Summary')
        ws3.merge_cells('A1:F1')
        ws3['A1'].value     = 'BRANCH-WISE SUMMARY'
        ws3['A1'].font      = Font(bold=True, size=13, color='1a56db')
        ws3['A1'].alignment = Alignment(horizontal='center')

        headers3 = [
            'Branch', 'Total', 'Placed',
            'Not Placed', 'Placement %', 'Applications'
        ]
        for col, h in enumerate(headers3, 1):
            apply_style(ws3.cell(row=2, column=col, value=h), header_style)

        set_column_widths(ws3, [30, 10, 10, 12, 14, 14])

        for row_num, (val, label) in enumerate(StudentProfile.BRANCH_CHOICES, 3):
            total  = StudentProfile.objects.filter(branch=val).count()
            placed = StudentProfile.objects.filter(branch=val, placement_status='placed').count()
            apps   = Application.objects.filter(student__branch=val).count()
            pct_b  = round((placed / total * 100) if total else 0, 1)
            for col, v in enumerate([label, total, placed, total - placed, f'{pct_b}%', apps], 1):
                cell = ws3.cell(row=row_num, column=col, value=v)
                cell.alignment = Alignment(horizontal='center')
                cell.border    = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'),
                )

        # ════════════════════════════
        #  SHEET 4: Company Summary
        # ════════════════════════════
        ws4 = wb.create_sheet('Company Summary')
        ws4.merge_cells('A1:G1')
        ws4['A1'].value     = 'COMPANY-WISE SUMMARY'
        ws4['A1'].font      = Font(bold=True, size=13, color='1a56db')
        ws4['A1'].alignment = Alignment(horizontal='center')

        headers4 = [
            'Company', 'Industry', 'Jobs',
            'Applications', 'Shortlisted', 'Selected', 'Rejected'
        ]
        for col, h in enumerate(headers4, 1):
            apply_style(ws4.cell(row=2, column=col, value=h), header_style)

        set_column_widths(ws4, [26, 24, 8, 14, 14, 12, 12])

        for row_num, company in enumerate(Company.objects.filter(is_active=True), 3):
            apps = Application.objects.filter(job__company=company)
            for col, v in enumerate([
                company.name,
                company.get_industry_display(),
                JobPost.objects.filter(company=company).count(),
                apps.count(),
                apps.filter(status='shortlisted').count(),
                apps.filter(status='selected').count(),
                apps.filter(status='rejected').count(),
            ], 1):
                cell = ws4.cell(row=row_num, column=col, value=v)
                cell.alignment = Alignment(horizontal='center')
                cell.border    = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'),
                )

        # ── Response ─────────────────────────────
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="full_placement_report.xlsx"'
        wb.save(response)
        return response