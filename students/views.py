import openpyxl
from django.contrib import messages
from django.http import HttpResponse, FileResponse, Http404
from django.shortcuts import get_object_or_404, redirect, render
from django.views.generic import DetailView, ListView, View
import os

from accounts.mixins import AdminRequiredMixin, StudentRequiredMixin
from accounts.models import User

from .forms import (
    AcademicRecordForm,
    SkillForm,
    StudentProfileForm,
)
from .models import AcademicRecord, Skill, StudentProfile


# ════════════════════════════════════════════════
#  ADMIN: STUDENT MANAGEMENT
# ════════════════════════════════════════════════

class StudentListView(AdminRequiredMixin, ListView):
    model               = StudentProfile
    template_name       = 'students/student_list.html'
    context_object_name = 'students'
    paginate_by         = 20

    def get_queryset(self):
        qs     = StudentProfile.objects.select_related('user').order_by('roll_number')
        branch = self.request.GET.get('branch')
        year   = self.request.GET.get('year')
        status = self.request.GET.get('status')
        search = self.request.GET.get('search')
        if branch:
            qs = qs.filter(branch=branch)
        if year:
            qs = qs.filter(year=year)
        if status:
            qs = qs.filter(placement_status=status)
        if search:
            qs = qs.filter(
                roll_number__icontains=search
            ) | qs.filter(
                user__full_name__icontains=search
            )
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['branch_choices'] = StudentProfile.BRANCH_CHOICES
        ctx['year_choices']   = StudentProfile.YEAR_CHOICES
        ctx['placement_status_choices'] = [
            ('not_placed',     'Not Placed'),
            ('placed',         'Placed'),
            ('higher_studies', 'Higher Studies'),
        ]
        return ctx


class StudentDetailView(AdminRequiredMixin, DetailView):
    model               = StudentProfile
    template_name       = 'students/student_detail.html'
    context_object_name = 'student'

    def get_queryset(self):
        return StudentProfile.objects.select_related('user').prefetch_related(
            'academic_records', 'skills'
        )


class AdminCreateStudentView(AdminRequiredMixin, View):
    template_name = 'students/student_form.html'

    def get(self, request):
        return render(request, self.template_name, {
            'form': StudentProfileForm()
        })

    def post(self, request):
        form = StudentProfileForm(request.POST, request.FILES)
        if form.is_valid():
            email     = form.cleaned_data['email']
            password  = form.cleaned_data['password']
            full_name = form.cleaned_data['full_name']

            if User.objects.filter(email=email).exists():
                messages.error(request, 'A user with this email already exists.')
                return render(request, self.template_name, {'form': form})

            user = User.objects.create_user(
                email=email,
                password=password,
                full_name=full_name,
                role='student'
            )
            profile      = form.save(commit=False)
            profile.user = user
            profile.save()

            # Send welcome email
            try:
                from applications.emails import send_welcome_email
                send_welcome_email(user, password)
            except Exception:
                pass

            messages.success(
                request,
                f'Student {full_name} created. Welcome email sent to {email}.'
            )
            return redirect('students:student_list')
        return render(request, self.template_name, {'form': form})


class UpdatePlacementStatusView(AdminRequiredMixin, View):

    def post(self, request, pk):
        student = get_object_or_404(StudentProfile, pk=pk)
        status  = request.POST.get('placement_status')
        if status in ['placed', 'not_placed', 'higher_studies']:
            student.placement_status = status
            student.save()
            messages.success(request, 'Placement status updated.')
        return redirect('students:student_detail', pk=pk)


# ════════════════════════════════════════════════
#  ADMIN: RESUME VIEWER
# ════════════════════════════════════════════════

class ResumeViewerView(AdminRequiredMixin, View):
    """View a student's resume inside the portal (PDF viewer)."""
    template_name = 'students/resume_viewer.html'

    def get(self, request, pk):
        student = get_object_or_404(
            StudentProfile.objects.select_related('user'), pk=pk
        )
        if not student.resume:
            messages.error(request, 'This student has not uploaded a resume.')
            return redirect('students:student_detail', pk=pk)

        return render(request, self.template_name, {
            'student':    student,
            'resume_url': student.resume.url,
            'is_pdf':     student.resume.name.lower().endswith('.pdf'),
        })


class ResumeDownloadView(AdminRequiredMixin, View):
    """Force-download a student's resume."""

    def get(self, request, pk):
        student = get_object_or_404(StudentProfile, pk=pk)
        if not student.resume:
            raise Http404("Resume not found.")

        file_path = student.resume.path
        if not os.path.exists(file_path):
            raise Http404("Resume file missing.")

        ext      = os.path.splitext(file_path)[1]
        filename = f"{student.roll_number}_resume{ext}"

        response = FileResponse(
            open(file_path, 'rb'),
            as_attachment=True,
            filename=filename
        )
        return response


class StudentResumeViewerView(StudentRequiredMixin, View):
    """Student views their own resume."""
    template_name = 'students/resume_viewer.html'

    def get(self, request):
        try:
            student = request.user.student_profile
        except StudentProfile.DoesNotExist:
            return redirect('students:create_profile')

        if not student.resume:
            messages.warning(
                request,
                'You have not uploaded a resume yet. Please update your profile.'
            )
            return redirect('students:update_profile')

        return render(request, self.template_name, {
            'student':    student,
            'resume_url': student.resume.url,
            'is_pdf':     student.resume.name.lower().endswith('.pdf'),
            'is_own':     True,
        })


# ════════════════════════════════════════════════
#  ADMIN: BULK UPLOAD
# ════════════════════════════════════════════════

class BulkUploadView(AdminRequiredMixin, View):
    template_name = 'students/bulk_upload.html'

    def get(self, request):
        return render(request, self.template_name, {
            'branch_choices': StudentProfile.BRANCH_CHOICES
        })

    def post(self, request):
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            messages.error(request, 'Please upload an Excel file.')
            return render(request, self.template_name, {
                'branch_choices': StudentProfile.BRANCH_CHOICES
            })

        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Only Excel files (.xlsx/.xls) are allowed.')
            return render(request, self.template_name, {
                'branch_choices': StudentProfile.BRANCH_CHOICES
            })

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
        except Exception:
            messages.error(request, 'Invalid Excel file. Please use the sample template.')
            return render(request, self.template_name, {
                'branch_choices': StudentProfile.BRANCH_CHOICES
            })

        success_count = 0
        error_rows    = []
        skipped_count = 0

        for row_num, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not any(row):
                continue
            try:
                (
                    full_name, email, password,
                    roll_number, branch, year,
                    cgpa, backlogs, phone, section
                ) = row[:10]

                if not all([full_name, email, password,
                            roll_number, branch, year, cgpa]):
                    error_rows.append(f'Row {row_num}: Missing required fields.')
                    continue

                if User.objects.filter(email=email).exists():
                    skipped_count += 1
                    continue

                if StudentProfile.objects.filter(roll_number=roll_number).exists():
                    skipped_count += 1
                    continue

                user = User.objects.create_user(
                    email=str(email).strip(),
                    password=str(password).strip(),
                    full_name=str(full_name).strip(),
                    role='student'
                )
                StudentProfile.objects.create(
                    user=user,
                    roll_number=str(roll_number).strip(),
                    branch=str(branch).strip().upper(),
                    year=int(year),
                    cgpa=float(cgpa),
                    backlogs=int(backlogs) if backlogs else 0,
                    phone=str(phone).strip() if phone else '',
                    section=str(section).strip() if section else '',
                )
                success_count += 1

            except Exception as e:
                error_rows.append(f'Row {row_num}: {str(e)}')
                continue

        if success_count:
            messages.success(
                request, f'✅ {success_count} student(s) uploaded successfully.'
            )
        if skipped_count:
            messages.warning(
                request, f'⚠️ {skipped_count} student(s) skipped (already exist).'
            )

        return render(request, self.template_name, {
            'success_count':  success_count,
            'skipped_count':  skipped_count,
            'error_rows':     error_rows,
            'branch_choices': StudentProfile.BRANCH_CHOICES,
        })


class DownloadSampleTemplateView(AdminRequiredMixin, View):

    def get(self, request):
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Students'

        headers = [
            'full_name', 'email', 'password',
            'roll_number', 'branch', 'year',
            'cgpa', 'backlogs', 'phone', 'section'
        ]
        for col, header in enumerate(headers, 1):
            cell           = ws.cell(row=1, column=col, value=header)
            cell.font      = Font(bold=True, color='FFFFFF')
            cell.fill      = PatternFill('solid', fgColor='1a56db')
            cell.alignment = Alignment(horizontal='center')

        sample_rows = [
            ['Ravi Kumar',   'ravi@college.edu',  'ravi@123',
             '1AJ23CS001', 'CSE', 3, 8.5, 0, '9876543210', 'A'],
            ['Priya Sharma', 'priya@college.edu', 'priya@123',
             '1AJ23CS002', 'CSE', 3, 7.8, 1, '9876543211', 'B'],
            ['Arjun Singh',  'arjun@college.edu', 'arjun@123',
             '1AJ23EC001', 'ECE', 4, 8.1, 0, '9876543212', 'A'],
            ['Neha Patel',   'neha@college.edu',  'neha@123',
             '1AJ23ME001', 'MECH', 3, 7.5, 0, '9876543213', 'C'],
            ['Karan Verma',  'karan@college.edu', 'karan@123',
             '1AJ23CV001', 'CIVIL', 4, 8.9, 0, '9876543214', 'A'],
        ]
        for row_num, row_data in enumerate(sample_rows, 2):
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)

        widths = [18, 24, 14, 16, 8, 6, 8, 10, 14, 10]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        ws['A8']      = 'Valid Branches: CSE, ECE, MECH, CIVIL, EEE, IT, AIDS, AIML'
        ws['A8'].font = Font(italic=True, color='6B7280')

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            'attachment; filename="sample_students_template.xlsx"'
        )
        wb.save(response)
        return response


# ════════════════════════════════════════════════
#  STUDENT: PROFILE MANAGEMENT
# ════════════════════════════════════════════════

class MyProfileView(StudentRequiredMixin, View):
    template_name = 'students/my_profile.html'

    def get(self, request):
        try:
            profile          = request.user.student_profile
            academic_records = profile.academic_records.all()
            skills           = profile.skills.all()
        except StudentProfile.DoesNotExist:
            return redirect('students:create_profile')
        return render(request, self.template_name, {
            'profile':          profile,
            'academic_records': academic_records,
            'skills':           skills,
        })


class CreateProfileView(StudentRequiredMixin, View):
    template_name = 'students/profile_form.html'

    def get(self, request):
        try:
            request.user.student_profile
            return redirect('students:my_profile')
        except StudentProfile.DoesNotExist:
            pass
        return render(request, self.template_name, {
            'form':  StudentProfileForm(),
            'title': 'Create Profile'
        })

    def post(self, request):
        form = StudentProfileForm(request.POST, request.FILES)
        if form.is_valid():
            profile      = form.save(commit=False)
            profile.user = request.user
            profile.save()
            messages.success(request, 'Profile created successfully!')
            return redirect('students:my_profile')
        return render(request, self.template_name, {
            'form':  form,
            'title': 'Create Profile'
        })


class UpdateProfileView(StudentRequiredMixin, View):
    template_name = 'students/profile_form.html'

    def get(self, request):
        try:
            profile = request.user.student_profile
        except StudentProfile.DoesNotExist:
            return redirect('students:create_profile')
        return render(request, self.template_name, {
            'form':  StudentProfileForm(instance=profile),
            'title': 'Edit Profile'
        })

    def post(self, request):
        try:
            profile = request.user.student_profile
        except StudentProfile.DoesNotExist:
            return redirect('students:create_profile')
        form = StudentProfileForm(
            request.POST, request.FILES, instance=profile
        )
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully!')
            return redirect('students:my_profile')
        return render(request, self.template_name, {
            'form':  form,
            'title': 'Edit Profile'
        })


class AddAcademicRecordView(StudentRequiredMixin, View):
    template_name = 'students/academic_form.html'

    def get(self, request):
        return render(request, self.template_name, {
            'form': AcademicRecordForm()
        })

    def post(self, request):
        form = AcademicRecordForm(request.POST)
        if form.is_valid():
            try:
                profile = request.user.student_profile
            except StudentProfile.DoesNotExist:
                return redirect('students:create_profile')
            record         = form.save(commit=False)
            record.student = profile
            record.save()
            messages.success(request, 'Academic record added.')
            return redirect('students:my_profile')
        return render(request, self.template_name, {'form': form})


class AddSkillView(StudentRequiredMixin, View):
    template_name = 'students/skill_form.html'

    def get(self, request):
        return render(request, self.template_name, {
            'form': SkillForm()
        })

    def post(self, request):
        form = SkillForm(request.POST)
        if form.is_valid():
            try:
                profile = request.user.student_profile
            except StudentProfile.DoesNotExist:
                return redirect('students:create_profile')
            skill         = form.save(commit=False)
            skill.student = profile
            skill.save()
            messages.success(request, 'Skill added.')
            return redirect('students:my_profile')
        return render(request, self.template_name, {'form': form})